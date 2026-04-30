import openpyxl
from openpyxl.styles import Font, PatternFill
import json
import os

def generate_excel_report(json_report_path, excel_path):
    """Generate Excel report from pytest JSON report"""
    with open(json_report_path, 'r') as f:
        data = json.load(f)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Results"
    
    # Headers
    headers = ["Test Name", "Status", "Duration", "Error"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    row = 2
    for test in data.get('tests', []):
        ws.cell(row=row, column=1).value = test.get('nodeid', '')
        outcome = test.get('outcome', '')
        ws.cell(row=row, column=2).value = outcome
        ws.cell(row=row, column=3).value = test.get('duration', 0)
        
        if outcome == 'failed':
            error = test.get('call', {}).get('crash', {}).get('message', '')
            ws.cell(row=row, column=4).value = error
            # Color failed rows
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        else:
            # Color passed rows
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        
        row += 1
    
    wb.save(excel_path)
    print(f"Excel report generated: {excel_path}")

if __name__ == "__main__":
    generate_excel_report("report.json", "report.xlsx")